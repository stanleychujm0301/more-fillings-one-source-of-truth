"""Excel 报告导出（P3 实现）— openpyxl + KPMG 蓝主题样式。

包含：核查总览（指标卡 + 严重度/类型分布图）、差异清单（配色分级 + 冻结 + 筛选 +
定位与取值 + 准则引用）、真实/预期差异、披露覆盖、提取预警、A/H 画像、证据定位。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from loguru import logger

from ahcc.report import _style as S
from ahcc.schemas import Job

# 复用 _style 中的 KPMG 配色常量（保留旧名以兼容外部引用）
KPMG_BLUE = S.KPMG_BLUE
KPMG_LIGHT_BLUE = S.KPMG_LIGHT_BLUE
KPMG_PURPLE = S.KPMG_PURPLE
_ILLEGAL_EXCEL_CHAR_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
_EXCEL_CELL_LIMIT = 32767
_NUMBER_FORMAT = "#,##0.00;(#,##0.00)"  # 会计括号负数，去掉刺眼 [Red]
_BODY_FONT = S.FONT_FAMILY  # 统一使用 Apple/金融风格主字体


def _clean_cell(value: Any) -> Any:
    """Remove PDF-extracted control characters that openpyxl rejects."""
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL_EXCEL_CHAR_RE.sub("", value)
    if len(cleaned) > _EXCEL_CELL_LIMIT:
        return cleaned[: _EXCEL_CELL_LIMIT - 3] + "..."
    return cleaned


def _append_row(ws, values: list[Any]) -> None:
    ws.append([_clean_cell(value) for value in values])


def _format_export_value(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return f"{value:,.0f}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


def _side_labels(job: Job | None = None) -> dict[str, str]:
    if job and getattr(job, "check_mode", "ah") == "h_bilingual":
        return {"A": "中文", "H": "英文"}
    return {"A": "A", "H": "H"}


def _diff_explanation_text(diff) -> str:
    explanation = getattr(diff, "diff_explanation", None)
    if not explanation:
        return diff.summary.best()
    parts = [explanation.headline, explanation.issue, explanation.location]
    return "\n".join(part for part in parts if part)


def _side_location_and_value(diff, side: str, labels: dict[str, str] | None = None) -> str:
    side_label = (labels or _side_labels()).get(side, side)
    explanation = getattr(diff, "diff_explanation", None)
    if explanation and explanation.items:
        lines: list[str] = []
        for item in explanation.items:
            page = item.a_page if side == "A" else item.h_page
            value = item.a_value if side == "A" else item.h_value
            snippet = item.a_snippet if side == "A" else item.h_snippet
            page_text = f"{side_label} 第{page}页" if page else f"{side_label} 未定位页码"
            label = item.label or "取值"
            value_text = _format_export_value(value)
            line = f"{page_text} | {label}: {value_text}"
            if snippet:
                line += f" | {snippet}"
            lines.append(line)
        return "\n".join(lines)

    evidence = [
        ev for ev in diff.evidence
        if (ev.side.value if hasattr(ev.side, "value") else ev.side) == side
    ]
    return "\n".join(
        f"{side_label} 第{ev.page}页 | {ev.snippet or ''}" if ev.page else f"{side_label} 未定位页码 | {ev.snippet or ''}"
        for ev in evidence
    )


def _review_hint(diff) -> str:
    explanation = getattr(diff, "diff_explanation", None)
    return explanation.review_hint if explanation and explanation.review_hint else ""


# ============================================================
# 样式工具
# ============================================================
def _visual_len(text: str) -> int:
    """估算显示宽度：CJK 字符按 2 计，取多行中最长一行。"""
    if not text:
        return 0
    longest = 0
    for line in str(text).split("\n"):
        width = sum(2 if ord(ch) > 0x2E80 else 1 for ch in line)
        longest = max(longest, width)
    return longest


def _auto_column_widths(ws, min_width: int = 9, max_width: int = 60, sample_rows: int = 300) -> None:
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        widest = 0
        for row in range(1, min(ws.max_row, sample_rows) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                widest = max(widest, _visual_len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, max(min_width, widest + 2))


def _style_header(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if ws.max_row < 1:
        return
    fill = PatternFill(start_color=S.HEADER_BG, end_color=S.HEADER_BG, fill_type="solid")
    bottom = Border(bottom=Side(style="medium", color=S.HEADER_BOTTOM))
    for cell in ws[1]:
        cell.font = Font(bold=True, color=S.HEADER_TEXT, name=_BODY_FONT, size=12)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bottom
    ws.row_dimensions[1].height = 38


def _apply_body_style(ws) -> None:
    """正文换行 + 顶部对齐 + 近白隔行 + 横向细线（去竖线，编辑式留白）。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    align = Alignment(vertical="top", wrap_text=True)
    font = Font(name=_BODY_FONT, size=10, color=S.INK)
    stripe = PatternFill(start_color=S.STRIPE, end_color=S.STRIPE, fill_type="solid")
    hair = Side(style="thin", color=S.HAIRLINE)
    border = Border(bottom=hair)
    for row in range(2, ws.max_row + 1):
        if not ws.row_dimensions[row].height:
            ws.row_dimensions[row].height = 27
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align
            cell.font = font
            cell.border = border
            if row % 2 == 0:
                cell.fill = stripe


def _finalize_sheet(ws, widths: dict[str, int] | None = None) -> None:
    """统一收尾：表头样式 + 正文样式 + 冻结首行 + 自动筛选 + 列宽。"""
    if ws.max_row < 1:
        return
    _style_header(ws)
    _apply_body_style(ws)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.zoomScaleNormal = 90
    ws.sheet_properties.tabColor = S.KPMG_BLUE
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    if widths:
        headers = [c.value for c in ws[1]]
        from openpyxl.utils import get_column_letter

        for idx, header in enumerate(headers, start=1):
            key = next((k for k in widths if header == k or (k.startswith("*") and str(header).endswith(k[1:]))), None)
            if key:
                ws.column_dimensions[get_column_letter(idx)].width = widths[key]
            else:
                # 未指定列回退到内容自适应
                widest = max(
                    (_visual_len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, min(ws.max_row, 300) + 1)),
                    default=0,
                )
                ws.column_dimensions[get_column_letter(idx)].width = min(45, max(9, widest + 2))
    else:
        _auto_column_widths(ws)
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.55
    ws.page_margins.bottom = 0.55
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _finalize_workbook(wb) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = S.KPMG_BLUE


# ============================================================
# 主入口
# ============================================================
def export_excel(job: Job, out_path: Path) -> None:
    """导出 Excel 差异报告。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "差异清单"

    side_labels = _side_labels(job)
    diff_headers = [
        "差异ID",
        "严重度",
        "分流",
        "差异类型",
        "主题",
        "差异说明",
        f"{side_labels['A']}定位与取值",
        f"{side_labels['H']}定位与取值",
        "A 股值",
        "H 股值",
        "差异",
        "证据页码",
        "AI 解读",
        "准则引用",
        "审阅提示",
        "审计师状态",
    ]
    _write_diff_sheet(ws, job.diffs, diff_headers, side_labels)

    ws_real = wb.create_sheet("真实差异")
    _write_diff_sheet(ws_real, [d for d in job.diffs if d.triage == "real"], diff_headers, side_labels)

    ws_expected = wb.create_sheet("预期差异")
    _write_diff_sheet(ws_expected, [d for d in job.diffs if d.triage == "expected"], diff_headers, side_labels)

    ws_coverage = wb.create_sheet("披露覆盖")
    _write_coverage_sheet(ws_coverage, job.coverage_items)

    ws_warnings = wb.create_sheet("提取预警")
    _write_warning_sheet(ws_warnings, job)

    ws_a = wb.create_sheet("A画像")
    _write_profile_sheet(ws_a, job.profile_a or {})

    ws_h = wb.create_sheet("H画像")
    _write_profile_sheet(ws_h, job.profile_h or {})

    ws_ev = wb.create_sheet("证据定位")
    _write_evidence_sheet(ws_ev, job.diffs)

    # 总览 sheet 放最前；但保持「差异清单」为 active（现有测试依赖 wb.active）
    import shutil

    out_path.parent.mkdir(parents=True, exist_ok=True)
    chart_dir = S.make_report_temp_dir(out_path.parent, "ahcc_xlsx_chart")
    try:
        ws_overview = wb.create_sheet("核查总览", 0)
        _write_overview_sheet(ws_overview, job, chart_dir)
        _finalize_workbook(wb)
        wb.active = wb.index(ws)

        wb.save(out_path)
    finally:
        shutil.rmtree(chart_dir, ignore_errors=True)
    logger.info(f"Excel 报告已导出：{out_path}")


# ============================================================
# 差异清单
# ============================================================
_DIFF_WIDTHS = {
    "差异ID": 11,
    "严重度": 9,
    "分流": 11,
    "差异类型": 12,
    "主题": 26,
    "差异说明": 48,
    "*定位与取值": 38,
    "A 股值": 16,
    "H 股值": 16,
    "差异": 15,
    "证据页码": 18,
    "AI 解读": 44,
    "准则引用": 32,
    "审阅提示": 32,
    "审计师状态": 12,
}


def _write_diff_sheet(ws, diffs, headers, side_labels: dict[str, str] | None = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    _append_row(ws, headers)

    sorted_diffs = sorted(diffs, key=lambda d: S.severity_rank(d.severity), reverse=True)
    for diff in sorted_diffs:
        pages = "; ".join(f"{e.side.value} P.{e.page}" for e in diff.evidence)
        ai_note = diff.standard_reasoning.rationale if diff.standard_reasoning else ""
        _append_row(ws, [
            diff.diff_id,
            S.severity_label_zh(diff.severity),
            S.triage_label_zh(diff.triage),
            S.diff_type_label_zh(diff.diff_type),
            diff.topic.best(),
            _diff_explanation_text(diff),
            _side_location_and_value(diff, "A", side_labels),
            _side_location_and_value(diff, "H", side_labels),
            diff.a_value,
            diff.h_value,
            diff.delta,
            pages,
            ai_note,
            S.standard_citation_text(diff),
            _review_hint(diff),
            diff.review_status.value,
        ])

    _finalize_sheet(ws, _DIFF_WIDTHS)

    # 列定位
    id_col = headers.index("差异ID") + 1
    sev_col = headers.index("严重度") + 1
    triage_col = headers.index("分流") + 1
    value_cols = [headers.index(h) + 1 for h in ("A 股值", "H 股值", "差异") if h in headers]

    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hair = Side(style="thin", color=S.HAIRLINE)
    chip_fill = PatternFill(start_color=S.REPORT_PANEL_SOFT, end_color=S.REPORT_PANEL_SOFT, fill_type="solid")
    chip_border = Border(
        left=Side(style="thin", color=S.REPORT_PANEL_BORDER),
        right=Side(style="thin", color=S.REPORT_PANEL_BORDER),
        top=Side(style="thin", color=S.REPORT_PANEL_BORDER),
        bottom=hair,
    )
    ws.sheet_view.zoomScale = 85
    ws.sheet_view.zoomScaleNormal = 85
    for offset, diff in enumerate(sorted_diffs):
        row = offset + 2
        sev_key = str(getattr(diff.severity, "value", diff.severity)).lower()
        ws.row_dimensions[row].height = 36

        # 严重度：文字着色（高级别暗红加粗，其余灰）——不再整列实色块
        scell = ws.cell(row=row, column=sev_col)
        scell.font = Font(
            name=_BODY_FONT, size=10,
            bold=S.severity_is_high(sev_key),
            color=S.severity_accent(sev_key),
        )
        scell.alignment = center
        scell.fill = chip_fill
        scell.border = chip_border

        # 分流：文字着色（real 暗红，其余灰）
        tri_key = str(getattr(diff.triage, "value", diff.triage)).lower()
        tcell = ws.cell(row=row, column=triage_col)
        tcell.font = Font(
            name=_BODY_FONT, size=10,
            bold=S.triage_is_real(tri_key),
            color=S.triage_accent(tri_key),
        )
        tcell.alignment = center
        tcell.fill = chip_fill
        tcell.border = chip_border

        # 左侧强调条：按严重度粗细/颜色变化
        width = S.severity_border_width(sev_key)
        if width:
            color = S.ALERT if S.severity_is_high(sev_key) else S.INK_SOFT
            idcell = ws.cell(row=row, column=id_col)
            idcell.border = Border(
                left=Side(style="medium" if width >= 2.0 else "thin", color=color),
                bottom=hair,
            )
            idcell.font = Font(name=_BODY_FONT, size=10, bold=True, color=S.INK)

        for vcol in value_cols:
            vcell = ws.cell(row=row, column=vcol)
            if isinstance(vcell.value, (int, float)):
                vcell.number_format = _NUMBER_FORMAT
            vcell.alignment = Alignment(horizontal="right", vertical="top")

    # 「差异」列数据条：克制灰阶，突出数值量级
    if "差异" in headers and ws.max_row >= 2:
        delta_col_idx = headers.index("差异") + 1
        delta_col = get_column_letter(delta_col_idx)
        ws.conditional_formatting.add(
            f"{delta_col}2:{delta_col}{ws.max_row}",
            DataBarRule(
                start_type="min", end_type="max",
                color=S.INK_SOFT, showValue=True, minLength=None, maxLength=None,
            ),
        )


# ============================================================
# 核查总览
# ============================================================
def _premium_title_block(
    ws,
    row: int,
    title: str,
    subtitle: str = "",
    conclusion: str = "",
    priority_note: str = "",
) -> int:
    """One-page hero 标题块：浅色 landing-page 首页 + 右侧优先级卡片。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    span_last = "I"
    hero_fill = PatternFill(start_color=S.HERO_WASH, end_color=S.HERO_WASH, fill_type="solid")
    wordmark_fill = PatternFill(start_color=S.REPORT_SURFACE, end_color=S.REPORT_SURFACE, fill_type="solid")
    panel_fill = PatternFill(start_color=S.REPORT_PANEL, end_color=S.REPORT_PANEL, fill_type="solid")
    accent = Side(style="thin", color=S.REPORT_SECTION_RULE)
    hair = Side(style="thin", color=S.REPORT_PANEL_BORDER)
    for rr in range(row, row + 8):
        for cc in range(1, 10):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = wordmark_fill if rr == row else hero_fill
            if rr == row + 7:
                cell.border = Border(bottom=accent)
            else:
                cell.border = Border()

    # eyebrow 字标
    ws.merge_cells(f"A{row}:{span_last}{row}")
    c = ws.cell(row=row, column=1, value=S.WORDMARK)
    c.font = Font(name=_BODY_FONT, size=9, bold=True, color=S.INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    # 大标题
    ws.merge_cells(f"A{row}:F{row + 2}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name=_BODY_FONT, size=24, bold=True, color=S.KPMG_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30
    ws.row_dimensions[row + 1].height = 30
    ws.row_dimensions[row + 2].height = 18

    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    c = ws.cell(row=row, column=7, value="ONE PAGE REVIEW")
    c.font = Font(name=_BODY_FONT, size=9, bold=True, color=S.KPMG_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="bottom")
    ws.merge_cells(start_row=row + 1, start_column=7, end_row=row + 2, end_column=9)
    c = ws.cell(row=row + 1, column=7, value=conclusion or "Evidence-led review")
    c.font = Font(name=_BODY_FONT, size=15, bold=True, color=S.ALERT if conclusion and "重点" in conclusion else S.KPMG_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 3

    # 副标题（公司名）
    if subtitle:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"项目名称 · {subtitle}")
        c.font = Font(name=_BODY_FONT, size=11, color=S.INK_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    c = ws.cell(row=row, column=7, value=priority_note or "Evidence-led review")
    c.font = Font(name=_BODY_FONT, size=9, color=S.INK_SOFT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # Right-side priority card: white surface floating on the light hero.
    for rr in range(row - 3, row + 1):
        for cc in range(7, 10):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = panel_fill
            cell.border = Border(
                left=hair if cc == 7 else Side(style=None),
                right=hair if cc == 9 else Side(style=None),
                top=hair if rr == row - 3 else Side(style=None),
                bottom=hair if rr == row else Side(style=None),
            )

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="Evidence-led disclosure consistency review")
    c.font = Font(name=_BODY_FONT, size=9, color=S.INK_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
    c = ws.cell(row=row, column=7, value="Executive Report")
    c.font = Font(name=_BODY_FONT, size=9, bold=True, color=S.KPMG_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

    row += 3
    ws.row_dimensions[row - 1].height = 5
    return row


def _write_context_strip(ws, start_row: int, items: list[tuple[str, object]]) -> int:
    """横向信息轨道，替代传统元数据表。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    groups = [(1, 2), (3, 4), (5, 6), (7, 9)]
    panel = PatternFill(start_color=S.REPORT_PANEL, end_color=S.REPORT_PANEL, fill_type="solid")
    hair = Side(style="thin", color=S.REPORT_PANEL_BORDER)
    accent = Side(style="thin", color=S.REPORT_SECTION_RULE)
    for (sc, ec), (label, value) in zip(groups, items):
        ws.merge_cells(start_row=start_row, start_column=sc, end_row=start_row, end_column=ec)
        ws.merge_cells(start_row=start_row + 1, start_column=sc, end_row=start_row + 1, end_column=ec)
        for rr in (start_row, start_row + 1):
            for cc in range(sc, ec + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = panel
                cell.border = Border(top=accent if rr == start_row else Side(style=None), bottom=hair)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row=start_row, column=sc, value=label).font = Font(name=_BODY_FONT, size=8.5, color=S.FOOTER_TEXT)
        ws.cell(row=start_row + 1, column=sc, value=_clean_cell(str(value))).font = Font(name=_BODY_FONT, size=9.5, bold=True, color=S.INK)
    ws.row_dimensions[start_row].height = 17
    ws.row_dimensions[start_row + 1].height = 23
    return start_row + 3


def _write_review_path(ws, start_row: int) -> int:
    """浅色三段审阅动线：把报告从“结果”导向“复核动作”。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    _section_header(ws, start_row, "审阅动线", span=9)
    card_row = start_row + 1
    note_row = start_row + 2
    groups = [
        (1, 3, "01 真实差异", "优先复核真实差异和重大项目"),
        (4, 6, "02 证据定位", "沿 A/H 页码、摘录和差异值追溯"),
        (7, 9, "03 人工复核", "形成审阅结论并更新状态"),
    ]
    panel = PatternFill(start_color=S.HERO_WASH, end_color=S.HERO_WASH, fill_type="solid")
    border = Border(bottom=Side(style="thin", color=S.REPORT_PANEL_BORDER))
    for sc, ec, title, note in groups:
        ws.merge_cells(start_row=card_row, start_column=sc, end_row=card_row, end_column=ec)
        ws.merge_cells(start_row=note_row, start_column=sc, end_row=note_row, end_column=ec)
        for row in (card_row, note_row):
            for col in range(sc, ec + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = panel
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=card_row, column=sc, value=title).font = Font(name=_BODY_FONT, size=10, bold=True, color=S.KPMG_BLUE)
        ws.cell(row=note_row, column=sc, value=note).font = Font(name=_BODY_FONT, size=8, color=S.INK_SOFT)
    ws.row_dimensions[card_row].height = 18
    ws.row_dimensions[note_row].height = 18
    return start_row + 3


def _write_kpi_cards(ws, start_row: int, metrics: list[tuple]) -> int:
    """合并单元格 KPI 卡片：2 行 × 3 列；每卡 = 标签 / 大数字 / 说明 + 顶部 accent。

    metrics: list of (label, value, note, is_alert)。返回下一可用行号。
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    panel = PatternFill(start_color=S.REPORT_PANEL, end_color=S.REPORT_PANEL, fill_type="solid")
    # 三张卡片三等分 A:C / D:F / G:I
    col_groups = [(1, 3), (4, 6), (7, 9)]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rows_per_card = 3
    gap_rows = 1

    for grp in range(2):  # 两行卡片
        base = start_row + grp * (rows_per_card + gap_rows)
        for ci, (sc, ec) in enumerate(col_groups):
            idx = grp * 3 + ci
            if idx >= len(metrics):
                break
            label, value, note, is_alert = metrics[idx]
            alert_on = bool(is_alert and value)
            accent = S.ALERT if alert_on else S.KPMG_BLUE

            # 三行：标签 / 数字 / 说明
            for dr in range(rows_per_card):
                ws.merge_cells(start_row=base + dr, start_column=sc, end_row=base + dr, end_column=ec)
                for col in range(sc, ec + 1):
                    ws.cell(row=base + dr, column=col).fill = panel

            lcell = ws.cell(row=base, column=sc, value=label)
            lcell.font = Font(name=_BODY_FONT, size=9, color=S.FOOTER_TEXT)
            lcell.alignment = center
            vcell = ws.cell(row=base + 1, column=sc, value=value)
            vcell.font = Font(name=_BODY_FONT, size=22, bold=False, color=S.ALERT if alert_on else S.INK)
            vcell.alignment = center
            ncell = ws.cell(row=base + 2, column=sc, value=note)
            ncell.font = Font(name=_BODY_FONT, size=8, color=S.FOOTER_TEXT)
            ncell.alignment = center

            # 顶部 accent line
            top = Side(style="medium" if alert_on else "thin", color=accent)
            for col in range(sc, ec + 1):
                cur = ws.cell(row=base, column=col)
                cur.border = Border(top=top)

            ws.row_dimensions[base].height = 16
            ws.row_dimensions[base + 1].height = 30
            ws.row_dimensions[base + 2].height = 16
        if grp == 0:
            ws.row_dimensions[base + rows_per_card].height = 8

    return start_row + 2 * (rows_per_card + gap_rows)


def _embed_distribution_charts(ws, job: Job, chart_dir: Path, anchor_row: int) -> bool:
    """渲染严重度/类型分布为 PNG 并嵌入；成功返回 True，失败 False（调用方回退原生图）。"""
    from openpyxl.drawing.image import Image as XLImage

    from ahcc.report._charts import donut_png, hbar_png

    sev = S.severity_distribution(job.diffs)
    typ = S.type_distribution(job.diffs)
    if not sev and not typ:
        return True  # 无差异：无需图表，视为完成

    embedded = False
    try:
        if sev:
            p = donut_png(sev, chart_dir / "sev.png", title="严重度分布")
            if p:
                img = XLImage(str(p))
                ratio = 360 / float(img.width)
                img.width = 360
                img.height = int(img.height * ratio)
                ws.add_image(img, f"A{anchor_row}")
                embedded = True
        if typ:
            p = hbar_png(typ, chart_dir / "typ.png", title="差异类型分布")
            if p:
                img = XLImage(str(p))
                ratio = 360 / float(img.width)
                img.width = 360
                img.height = int(img.height * ratio)
                ws.add_image(img, f"F{anchor_row}")
                embedded = True
    except Exception as exc:  # noqa: BLE001
        logger.warning(f"Excel 分布图嵌入失败，将回退原生图表：{exc}")
        return False
    return embedded


def _paint_chart_stage(ws, anchor_row: int) -> None:
    """Paint light dashboard panels behind embedded chart images."""
    from openpyxl.styles import Border, PatternFill, Side

    panel = PatternFill(start_color=S.REPORT_PANEL, end_color=S.REPORT_PANEL, fill_type="solid")
    hair = Side(style="thin", color=S.REPORT_PANEL_BORDER)
    for start_col, end_col in ((1, 4), (6, 9)):
        for rr in range(anchor_row, anchor_row + 13):
            ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 0, 18)
            for cc in range(start_col, end_col + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = panel
                cell.border = Border(
                    left=hair if cc == start_col else Side(style=None),
                    right=hair if cc == end_col else Side(style=None),
                    top=hair if rr == anchor_row else Side(style=None),
                    bottom=hair if rr == anchor_row + 12 else Side(style=None),
                )


def _write_overview_sheet(ws, job: Job, chart_dir: Path | None = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    summary = job.comparison_summary or {}
    bilingual = getattr(job, "check_mode", "ah") == "h_bilingual"
    title = "H 股中英文报告一致性核查报告" if bilingual else "A+H 股年报数据一致性核查报告"
    company = job.company_name or "—"
    real_count = summary.get("real_diff_count", sum(1 for d in job.diffs if d.triage == "real"))
    unresolved_count = summary.get("unresolved_diff_count", sum(1 for d in job.diffs if d.triage == "unresolved"))
    conclusion = S.conclusion_label(int(real_count or 0), int(unresolved_count or 0))

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = S.KPMG_BLUE
    ws.freeze_panes = "A2"

    # 列宽：三等分卡片 + 适度留白
    for col, w in (("A", 13), ("B", 13), ("C", 13), ("D", 13), ("E", 13),
                   ("F", 13), ("G", 13), ("H", 13), ("I", 13)):
        ws.column_dimensions[col].width = w

    # —— premium 标题块 ——
    row = _premium_title_block(
        ws,
        1,
        title,
        company,
        conclusion,
        f"真实差异 {real_count} 项 · 待判断 {unresolved_count} 项",
    )

    # —— context strip ——
    gen_time = S.format_beijing_datetime(job.finished_at or job.started_at)
    dur = S.format_duration(job.duration_seconds)
    context_items = [
        ("项目名称", company),
        ("核查模式", summary.get("mode_label", S.check_mode_label(getattr(job, "check_mode", "ah")))),
        ("生成时间", gen_time),
        ("核查耗时", dur),
    ]
    row = _write_context_strip(ws, row, context_items)

    # —— 执行摘要 ——
    _section_header(ws, row, "执行摘要", span=9)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value=conclusion).font = Font(name=_BODY_FONT, size=14, bold=True, color=S.ALERT if real_count else S.KPMG_BLUE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=row, column=4, value=f"审阅提示：优先复核真实差异、待判断事项及证据定位；完整明细见「差异清单」。").font = Font(name=_BODY_FONT, size=9, color=S.INK_SOFT)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fill = PatternFill(start_color=S.REPORT_PANEL, end_color=S.REPORT_PANEL, fill_type="solid")
    hair = Side(style="thin", color=S.REPORT_PANEL_BORDER)
    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = Border(bottom=hair)
    ws.row_dimensions[row].height = 28

    # —— 审阅动线 ——
    row += 1
    row = _write_review_path(ws, row)

    # —— KPI 卡片 ——
    metrics = [
        ("差异总数", summary.get("total_diff_count", len(job.diffs)), "全部差异条数", False),
        ("真实差异", real_count, "需重点关注", True),
        ("预期差异", summary.get("expected_diff_count", sum(1 for d in job.diffs if d.triage == "expected")), "可解释差异", False),
        ("待判断", unresolved_count, "需人工确认", False),
        ("披露覆盖", summary.get("coverage_count", len(job.coverage_items)), "匹配 / 单边项", False),
        ("核查耗时", dur, f"生成时间 {gen_time}", False),
    ]
    row = _write_kpi_cards(ws, row, metrics)

    # 双语模式补充指标
    if bilingual:
        row += 1
        _section_header(ws, row, "中英文核查指标", span=3)
        row += 1

        def pct(v):
            return f"{(v or 0) * 100:.1f}%"

        sem_total = summary.get("semantic_total_pairs", 0)
        sem_cell = (
            f"{pct(summary.get('semantic_coverage', 0))}（{summary.get('semantic_reviewed_pairs', 0)}/{sem_total} 对）"
            if sem_total else "未启用"
        )
        bi_rows = [
            ("翻译覆盖率", pct(summary.get("translation_coverage")), "已配对中文段落 / 中文总段落"),
            ("表格覆盖率", pct(summary.get("table_coverage")), "已配对中文表格 / 中文总表格"),
            ("LLM 翻译审查覆盖", sem_cell, "strict 模式下分批审查的段落覆盖比"),
            ("跨币种核对一致", summary.get("cross_currency_matched", 0), "人民币/港币换算后一致项"),
            ("跨币种不一致", summary.get("cross_currency_mismatch", 0), "换算后仍不一致，已降级"),
            ("币种不可识别", summary.get("currency_ambiguous", 0), "一侧币种缺失，需人工复核"),
        ]
        for label, value, note in bi_rows:
            ws.cell(row=row, column=1, value=label).font = Font(name=_BODY_FONT, size=10, bold=True)
            ws.cell(row=row, column=2, value=_clean_cell(str(value))).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=_clean_cell(str(note))).font = Font(name=_BODY_FONT, size=9, color=S.TEXT_MUTED)
            row += 1

    # —— 分布图 ——
    row += 1
    _section_header(ws, row, "分布概览", span=9)
    row += 1
    _paint_chart_stage(ws, row)
    embedded = False
    if chart_dir is not None:
        embedded = _embed_distribution_charts(ws, job, chart_dir, row)
    if not embedded:
        _write_native_charts(ws, job, row)


def _write_native_charts(ws, job: Job, anchor_row: int) -> None:
    """回退：openpyxl 原生 BarChart（仅当 PNG 渲染不可用时）。"""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    sev_dist = S.severity_distribution(job.diffs)
    type_dist = S.type_distribution(job.diffs)

    data_top = anchor_row
    # 数据写到远右列（L/M），避免与版面冲突
    sev_col, cnt_col = 12, 13
    ws.cell(row=data_top, column=sev_col, value="严重度").font = Font(name=_BODY_FONT, bold=True)
    ws.cell(row=data_top, column=cnt_col, value="数量").font = Font(name=_BODY_FONT, bold=True)
    r = data_top + 1
    for label, count in (sev_dist or {"（无差异）": 0}).items():
        ws.cell(row=r, column=sev_col, value=label)
        ws.cell(row=r, column=cnt_col, value=count)
        r += 1
    sev_last = r - 1

    if sev_dist:
        chart = BarChart()
        chart.type = "col"
        chart.title = "严重度分布"
        chart.height = 8
        chart.width = 13
        chart.legend = None
        data = Reference(ws, min_col=cnt_col, min_row=data_top, max_row=sev_last)
        cats = Reference(ws, min_col=sev_col, min_row=data_top + 1, max_row=sev_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _style_mono_chart(chart, sev_last - data_top)
        ws.add_chart(chart, f"A{anchor_row}")

    type_top = sev_last + 3
    ws.cell(row=type_top, column=sev_col, value="差异类型").font = Font(name=_BODY_FONT, bold=True)
    ws.cell(row=type_top, column=cnt_col, value="数量").font = Font(name=_BODY_FONT, bold=True)
    r = type_top + 1
    for label, count in (type_dist or {"（无差异）": 0}).items():
        ws.cell(row=r, column=sev_col, value=label)
        ws.cell(row=r, column=cnt_col, value=count)
        r += 1
    type_last = r - 1

    if type_dist:
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "差异类型分布"
        chart2.height = 8
        chart2.width = 13
        chart2.legend = None
        data2 = Reference(ws, min_col=cnt_col, min_row=type_top, max_row=type_last)
        cats2 = Reference(ws, min_col=sev_col, min_row=type_top + 1, max_row=type_last)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        _style_mono_chart(chart2, type_last - type_top)
        ws.add_chart(chart2, f"F{anchor_row}")


def _style_mono_chart(chart, n_points: int) -> None:
    """把 BarChart 收敛为海军蓝单色阶：去网格线、去图例、按数据点上深浅蓝。"""
    try:
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties

        chart.varyColors = False
        chart.legend = None
        chart.gapWidth = 60
        if getattr(chart, "y_axis", None) is not None:
            chart.y_axis.majorGridlines = None
        if not chart.series:
            return
        series = chart.series[0]
        series.graphicalProperties = GraphicalProperties(solidFill=S.MONO_RAMP[0])
        for i in range(max(0, n_points)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties = GraphicalProperties(solidFill=S.mono_color(i))
            series.data_points.append(pt)
    except Exception as exc:  # 图表着色失败不阻断报告
        logger.warning(f"Excel 单色图着色失败：{exc}")


def _section_header(ws, row: int, text: str, span: int = 3) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=_BODY_FONT, size=12, bold=True, color=S.INK)
    cell.fill = PatternFill(start_color=S.REPORT_SECTION_FILL, end_color=S.REPORT_SECTION_FILL, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # 浅底 + 墨色字 + 底部细线（去掉高饱和紫色块）
    underline = Side(style="thin", color=S.REPORT_SECTION_RULE)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).border = Border(bottom=underline)
    ws.row_dimensions[row].height = 22


# ============================================================
# 其余 sheet
# ============================================================
def _write_profile_sheet(ws, profile: dict) -> None:
    _append_row(ws, ["范围", "键/主题", "名称", "值/字数", "单位/层级", "页码/范围", "证据/摘要"])
    _append_row(ws, ["summary", "doc_id", profile.get("doc_id", ""), "", "", "", ""])
    _append_row(ws, ["summary", "total_pages", profile.get("total_pages", ""), "", "", "", ""])
    _append_row(ws, ["summary", "metric_occurrences", profile.get("metric_occurrences", ""), "", "", "", ""])
    _append_row(ws, ["summary", "narrative_blocks", profile.get("narrative_blocks", ""), "", "", "", ""])
    audit = profile.get("extraction_audit") or {}
    if audit:
        _append_row(ws, ["audit", "scanned_pages", len(audit.get("scanned_pages") or []), "", "", "", ""])
        _append_row(ws, ["audit", "coverage_ratio", audit.get("coverage_ratio", ""), "", "", "", ""])
        _append_row(ws, ["audit", "blank_pages", len(audit.get("blank_pages") or []), "", "", "", ""])
        _append_row(ws, ["audit", "ocr_pages", len(audit.get("ocr_pages") or []), "", "", "", ""])
        _append_row(ws, ["audit", "table_pages", len(audit.get("table_pages") or []), "", "", "", ""])
        _append_row(ws, ["audit", "warnings", " | ".join(audit.get("warnings") or []), "", "", "", ""])

    for item in profile.get("metrics", []):
        ev = item.get("evidence", {}) or {}
        name = item.get("name", {}) or {}
        _append_row(ws, [
            "metric",
            item.get("canonical_key", ""),
            name.get("zh") or name.get("en") or "",
            item.get("value"),
            item.get("unit") or item.get("currency") or "",
            item.get("page"),
            ev.get("snippet", ""),
        ])

    for item in profile.get("narratives", []):
        _append_row(ws, [
            "narrative",
            item.get("topic_key", ""),
            item.get("topic_label", ""),
            item.get("word_count"),
            item.get("detail_level", ""),
            str(item.get("page_range", "")),
            item.get("summary", ""),
        ])

    _finalize_sheet(ws, {"范围": 10, "键/主题": 22, "名称": 20, "值/字数": 14, "单位/层级": 14, "页码/范围": 14, "证据/摘要": 60})


def _write_coverage_sheet(ws, items) -> None:
    _append_row(ws, ["覆盖ID", "状态", "类别", "主题", "Key", "A页码", "H页码", "匹配置信度", "说明"])
    for item in items:
        _append_row(ws, [
            item.coverage_id,
            item.status,
            item.category,
            item.topic.best(),
            item.canonical_key or "",
            ", ".join(str(p) for p in item.a_pages),
            ", ".join(str(p) for p in item.h_pages),
            item.match_confidence,
            item.note,
        ])
    _finalize_sheet(ws, {"覆盖ID": 12, "状态": 10, "类别": 12, "主题": 26, "Key": 20, "A页码": 12, "H页码": 12, "匹配置信度": 12, "说明": 50})


def _write_warning_sheet(ws, job: Job) -> None:
    _append_row(ws, ["侧", "预警标识", "类别", "严重性", "阻断核心核查", "说明", "总页数", "扫描页数", "覆盖率", "缺失页", "空白页", "OCR页", "表格页"])
    warnings = (job.comparison_summary or {}).get("warnings") or []
    if warnings:
        for item in warnings:
            _append_row(ws, [
                item.get("side", ""),
                item.get("flag", ""),
                item.get("category", ""),
                item.get("severity", ""),
                "是" if item.get("blocking") else "否",
                item.get("message", ""),
                item.get("total_pages", 0),
                item.get("scanned_pages", 0),
                item.get("coverage_ratio", 0),
                item.get("missing_pages", 0),
                item.get("blank_pages", 0),
                item.get("ocr_pages", 0),
                item.get("table_pages", 0),
            ])
    else:
        for side, profile in (("A", job.profile_a or {}), ("H", job.profile_h or {})):
            audit = profile.get("extraction_audit") or {}
            if not audit:
                continue
            _append_row(ws, [
                side, "", "", "", "否", "No extraction warnings.",
                audit.get("total_pages", 0),
                len(audit.get("scanned_pages") or []),
                audit.get("coverage_ratio", 0),
                len(audit.get("missing_pages") or []),
                len(audit.get("blank_pages") or []),
                len(audit.get("ocr_pages") or []),
                len(audit.get("table_pages") or []),
            ])
    _finalize_sheet(ws, {"说明": 50})


def _write_evidence_sheet(ws, diffs) -> None:
    _append_row(ws, ["差异ID", "分流", "差异类型", "侧", "页码", "章节", "坐标(bbox)", "原文片段"])
    for diff in diffs:
        for ev in diff.evidence:
            _append_row(ws, [
                diff.diff_id,
                S.triage_label_zh(diff.triage),
                S.diff_type_label_zh(diff.diff_type),
                ev.side.value,
                ev.page,
                ev.section or "",
                _format_bbox(ev.bbox),
                ev.snippet,
            ])
    _finalize_sheet(ws, {"差异ID": 12, "分流": 11, "差异类型": 12, "侧": 6, "页码": 8, "章节": 22, "坐标(bbox)": 22, "原文片段": 60})


def _format_bbox(bbox) -> str:
    if not bbox:
        return ""
    try:
        return "(" + ", ".join(f"{float(v):.0f}" for v in bbox) + ")"
    except Exception:
        return str(bbox)
