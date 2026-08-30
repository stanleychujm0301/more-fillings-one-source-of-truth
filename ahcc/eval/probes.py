"""无标注度量：在没有标准答案的真实 A/H 样本上度量误报。

背景
----
7 组真实 A/H 年报（平安/电信/中芯/光大证券/光大银行/国泰海通/申万宏源）没有标准答案，
旧的 `scripts/eval_samples.py` 在找不到答案文件时直接 `return None` 把整个 job 结果丢弃 ——
**用户抱怨的误报全部发生在这 7 组上，而工具链对它们零观测。**

本模块提供两条不需要人工标注就能跑的度量通道：

1. `self_consistency_report` —— 把同一份 PDF 同时当 A 和 H 传入。
   两侧完全相同，因此**任何跨报告差异都是纯误报**，可直接定位到 rule_id。
   这是最便宜、最硬的 FP 探针，应进默认 pytest。

2. `export_fp_workbook` —— 真实 A/H 年报的同一指标本应一致，因此每一条
   `triage=="real" and scope==CROSS_REPORT` 的差异都是待验条目。全量条数即
   **FP 上界**；按 rule_id × triage × severity 分桶抽检，确认率给出 FP 率的
   Wilson 置信区间。
"""

from __future__ import annotations

import math
import random
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from ahcc.schemas import Diff, DiffScope


def _norm(value: Any) -> str:
    return str(getattr(value, "value", value) or "").strip().lower()


def _scope_of(diff: Diff) -> str:
    scope = getattr(diff, "diff_scope", DiffScope.CROSS_REPORT)
    return _norm(scope) or DiffScope.CROSS_REPORT.value


# ============================================================
# 探针 1：A=H 自一致性
# ============================================================

@dataclass
class SelfConsistencyReport:
    """同一份 PDF 自配对的结果。cross_report_count 必须为 0。"""

    pair_id: str
    total_diffs: int
    cross_report_diffs: list[Diff] = field(default_factory=list)
    by_rule_id: dict[str, int] = field(default_factory=dict)

    @property
    def cross_report_count(self) -> int:
        return len(self.cross_report_diffs)

    @property
    def passed(self) -> bool:
        return self.cross_report_count == 0

    def summary_line(self) -> str:
        status = "PASS" if self.passed else "FAIL"
        detail = ""
        if self.by_rule_id:
            top = sorted(self.by_rule_id.items(), key=lambda kv: -kv[1])[:5]
            detail = "  " + " ".join(f"{k}={v}" for k, v in top)
        return (
            f"[自一致性探针 {self.pair_id}] {status} "
            f"跨报告差异 {self.cross_report_count} / 总差异 {self.total_diffs}{detail}"
        )


def self_consistency_report(diffs: Iterable[Diff], *, pair_id: str = "") -> SelfConsistencyReport:
    """分析 A=H 自配对的输出。

    A 与 H 是同一份文件，任何 `diff_scope == cross_report` 的差异在定义上都是误报。
    单侧内部差异（a_internal / h_internal）是合法的 —— 一份报告确实可能自身不自洽 ——
    但它们会在两侧各出现一次，属预期行为，不计入探针。
    """
    diff_list = list(diffs)
    cross = [d for d in diff_list if _scope_of(d) == DiffScope.CROSS_REPORT.value]
    by_rule: dict[str, int] = defaultdict(int)
    for d in cross:
        by_rule[d.rule_id or "__unlabeled__"] += 1
    return SelfConsistencyReport(
        pair_id=pair_id,
        total_diffs=len(diff_list),
        cross_report_diffs=cross,
        by_rule_id=dict(by_rule),
    )


# ============================================================
# 探针 2：FP 上界 + 分桶抽检
# ============================================================

def wilson_interval(hits: int, total: int, z: float = 1.96) -> tuple[float, float]:
    """Wilson score 区间 —— 小样本抽检下比正态近似可靠得多。

    返回 (下界, 上界)。total==0 时返回 (0.0, 1.0)（完全未知）。
    """
    if total <= 0:
        return 0.0, 1.0
    p = hits / total
    denom = 1 + z * z / total
    centre = p + z * z / (2 * total)
    margin = z * math.sqrt(p * (1 - p) / total + z * z / (4 * total * total))
    lo = (centre - margin) / denom
    hi = (centre + margin) / denom
    return max(0.0, lo), min(1.0, hi)


@dataclass
class FpBucket:
    rule_id: str
    triage: str
    severity: str
    scope: str
    diffs: list[Diff] = field(default_factory=list)

    @property
    def key(self) -> tuple[str, str, str, str]:
        return (self.rule_id, self.triage, self.severity, self.scope)


def bucket_diffs(diffs: Iterable[Diff]) -> list[FpBucket]:
    """按 rule_id × triage × severity × scope 分桶。"""
    buckets: dict[tuple[str, str, str, str], FpBucket] = {}
    for d in diffs:
        key = (
            d.rule_id or "__unlabeled__",
            _norm(getattr(d, "triage", "")),
            _norm(d.severity),
            _scope_of(d),
        )
        bucket = buckets.get(key)
        if bucket is None:
            bucket = FpBucket(rule_id=key[0], triage=key[1], severity=key[2], scope=key[3])
            buckets[key] = bucket
        bucket.diffs.append(d)
    return sorted(buckets.values(), key=lambda b: (-len(b.diffs), b.key))


def fp_upper_bound(diffs: Iterable[Diff]) -> dict[str, int]:
    """真实 A/H 年报上的 FP 上界。

    同一家公司的 A 股与 H 股年报，同一指标本应一致。因此每一条被报为
    「真实的跨报告差异」都是待验条目 —— 在人工确认之前，它们的**全量条数**
    就是 FP 的上界（上界，因为其中确实可能有真差异）。
    """
    counts = {"cross_report_real": 0, "cross_report_unresolved": 0, "cross_report_expected": 0, "internal": 0}
    for d in diffs:
        scope = _scope_of(d)
        if scope != DiffScope.CROSS_REPORT.value:
            counts["internal"] += 1
            continue
        triage = _norm(getattr(d, "triage", "")) or "real"
        key = f"cross_report_{triage}"
        counts[key] = counts.get(key, 0) + 1
    return counts


def _page_str(diff: Diff) -> str:
    a = h = None
    for ev in diff.evidence:
        side = _norm(ev.side)
        if side == "a" and a is None:
            a = ev.page
        elif side == "h" and h is None:
            h = ev.page
    return f"A{a or '-'}/H{h or '-'}"


def export_fp_workbook(
    pair_id: str,
    diffs: Iterable[Diff],
    out_path: str | Path,
    *,
    sample_per_bucket: int = 20,
    seed: int = 0,
) -> dict[str, int]:
    """导出全量差异分桶 + 人工抽检工作簿。

    产出三张表：
    - 「FP上界」    —— 分档计数，直接就是 FP 上界
    - 「分桶统计」  —— 每个 rule_id × triage × severity × scope 的条数与抽检量
    - 「人工抽检」  —— 每桶随机抽 N 条，留空「是否真差异」列供人工回填

    回填后用 `wilson_interval` 就能给出该桶 FP 率的区间估计。
    """
    from openpyxl import Workbook

    diff_list = list(diffs)
    rng = random.Random(seed)
    buckets = bucket_diffs(diff_list)
    bounds = fp_upper_bound(diff_list)

    wb = Workbook()

    ws = wb.active
    ws.title = "FP上界"
    ws.append(["项目", "条数", "说明"])
    ws.append(["样本对", pair_id, ""])
    ws.append(["总差异", len(diff_list), ""])
    ws.append([
        "跨报告 · real", bounds.get("cross_report_real", 0),
        "FP 上界：A/H 同一指标本应一致，每条都需人工确认",
    ])
    ws.append(["跨报告 · unresolved", bounds.get("cross_report_unresolved", 0), "待人工判定"])
    ws.append(["跨报告 · expected", bounds.get("cross_report_expected", 0), "系统自称可解释，需抽检确认抑制是否正确"])
    ws.append(["单侧内部差异", bounds.get("internal", 0), "报告自身不自洽，不属于跨报告 FP 口径"])

    ws_b = wb.create_sheet("分桶统计")
    ws_b.append(["rule_id", "triage", "severity", "scope", "条数", "抽检量"])
    for b in buckets:
        ws_b.append([
            b.rule_id, b.triage, b.severity, b.scope,
            len(b.diffs), min(sample_per_bucket, len(b.diffs)),
        ])

    ws_s = wb.create_sheet("人工抽检")
    ws_s.append([
        "rule_id", "triage", "severity", "scope", "差异ID", "主题",
        "A值", "H值", "定位", "差异说明", "是否真差异(Y/N)", "备注",
    ])
    for b in buckets:
        picked = b.diffs if len(b.diffs) <= sample_per_bucket else rng.sample(b.diffs, sample_per_bucket)
        for d in picked:
            ws_s.append([
                b.rule_id, b.triage, b.severity, b.scope, d.diff_id,
                d.topic.best() if d.topic else "",
                d.a_value, d.h_value, _page_str(d),
                (d.summary.best() if d.summary else "")[:300],
                "", "",
            ])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return bounds


def summarize_sampling(rows: Iterable[tuple[str, bool]]) -> dict[str, dict[str, Any]]:
    """把人工回填的抽检结果汇总成每个 rule_id 的确认率与 Wilson 区间。

    `rows` 为 [(rule_id, 是否真差异), ...]。
    """
    stats: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for rule_id, is_real in rows:
        bucket = stats[rule_id]
        bucket[1] += 1
        if is_real:
            bucket[0] += 1
    out: dict[str, dict[str, Any]] = {}
    for rule_id, (hits, total) in stats.items():
        lo, hi = wilson_interval(hits, total)
        out[rule_id] = {
            "confirmed": hits,
            "sampled": total,
            "precision": round(hits / total, 4) if total else 0.0,
            "wilson_low": round(lo, 4),
            "wilson_high": round(hi, 4),
            "fp_rate_high": round(1 - lo, 4),
        }
    return out
