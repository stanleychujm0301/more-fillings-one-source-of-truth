"""答案对比核心逻辑：系统检出 Diff 与预期答案 ExpectedDiff 的匹配与指标计算。

设计原则（2026-08 重建，见 plan Phase 0.1）
------------------------------------------------
旧实现有三个会系统性虚高指标的问题，已全部修掉：

1. **命中判定曾把 `evidence.snippet` 拼进检索串再做无边界子串 `in`。**
   snippet 是 PDF 里的一整行原文，包含该行所有数字，于是"同页任意一条无关差异"
   都会命中。现在命中只看 diff **自己声称的值**（a_value/h_value/summary），
   snippet 只用于人工复核展示；且比较改为**数值 token 精确相等**，
   杜绝 `1,126,411` 命中 `126,411`。

2. **只命中 `original_value`（未被篡改的正确值）曾算 prefix 命中。**
   "系统报出了正确值"不等于"系统发现了篡改"。现在这类降级为 `weak`，
   单列统计，不计入 `hit_count`。

3. **命中不看 triage/severity。** 一条被自己降级为 expected/INFO、
   产品界面根本不展示的 diff 也算召回。现在召回按**用户可见性**定义
   （triage == "real" 且 severity ≥ medium）；查到了却被自己压掉的，
   单列 `detected_but_suppressed_count` —— 这个数字才是当前 FN 的主战场。

另外：贪心分配改为全局最优二分匹配（指标不再随答案行顺序波动）；
页码容差由模块常量统一控制；不再拿 H 侧页码去比 A 侧答案页。

匹配级别（按优先级）
- exact  — 答案的「错误值」与「原始值」都出现在 diff 声称的值里，或 rule_id 精确相等
- value  — 答案的「错误值」出现在 diff 声称的值里（方向正确）
- rule   — rule_id 互为前缀
- fuzzy  — 主题 canonical_key 有交集 + 页码接近（**不计入命中**，转人工确认）
- weak   — 只命中「原始值」（**不计入命中**）

指标
- 召回率 = 严格命中数 / ExpectedDiff 总数
- 精确率 = 严格命中数 / (严格命中数 + hard_FP)
- hard_FP     = 未命中任何预期且 triage == "real"     —— 真正会被用户当成错误的误报
- soft_FP     = 未命中任何预期且 triage == "unresolved" —— 待人工判定，成本较低
- suppressed  = 未命中任何预期且 triage == "expected"   —— 系统自称"可解释"，需抽检确认抑制是否正确
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ahcc.align.glossary import glossary, to_simplified
from ahcc.schemas import Diff, ReportSide

# 页码容差：统一在此控制，不要在函数体里再写字面量。
_PAGE_TOLERANCE = 3          # rule_id / fuzzy 路径
_VALUE_PAGE_TOLERANCE = 2    # 官方错误清单的值匹配路径（同一处篡改不会跨多页）

# 产品实际会呈现为"问题"的档位。召回率只认这些。
_VISIBLE_SEVERITIES = {"critical", "high", "medium"}

# 计入 hit_count 的匹配级别。fuzzy/weak 明确排除。
_STRICT_LEVELS = ("exact", "value", "rule")
_LEVEL_RANK = {"exact": 4, "value": 3, "rule": 2, "fuzzy": 1, "weak": 0}

_NUMBER_TOKEN_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")


@dataclass
class ExpectedDiff:
    pair_id: str = ""
    company: str = ""
    expected_rule_id: str = ""
    topic: str = ""
    expected_severity: str = ""
    a_page: int | None = None
    h_page: int | None = None
    note: str = ""
    # 官方错误清单格式专属字段（序号/PDF页码/描述/原始数字/错误数字）
    page: int | None = None
    original_value: str = ""
    tampered_value: str = ""
    description: str = ""


@dataclass
class MatchResult:
    expected: ExpectedDiff
    matched_diff: Diff | None = None
    match_level: str = "missed"  # exact / value / rule / fuzzy / weak / missed
    reason: str = ""
    # 被匹配上、但因 triage/severity 不达标而在产品里不可见的 diff。
    # 这类计入 detected_but_suppressed，不计入 hit。
    suppressed_diff: Diff | None = None
    suppressed_level: str = ""


@dataclass
class EvalReport:
    pair_id: str
    expected_count: int
    detected_count: int
    hit_count: int
    false_positive_count: int  # == hard_fp_count，保留旧字段名以兼容既有调用方
    recall: float
    precision: float
    matches: list[MatchResult] = field(default_factory=list)
    unmatched_diffs: list[Diff] = field(default_factory=list)
    # --- 重建后新增的分项指标 ---
    hard_fp_count: int = 0
    soft_fp_count: int = 0
    suppressed_count: int = 0
    needs_manual_review_count: int = 0
    weak_count: int = 0
    detected_but_suppressed_count: int = 0
    visible_diff_count: int = 0
    recall_by_rule_id: dict[str, tuple[int, int]] = field(default_factory=dict)


# ============================================================
# 答案加载（未改动逻辑，仅保持原样）
# ============================================================

def load_answer_key(path: Path) -> list[ExpectedDiff]:
    """读取预期答案 Excel。自动嗅探表头格式：

    - 官方错误清单格式（序号/PDF页码/描述/原始数字/错误数字/…）→ load_official_answer_key
    - 内部格式（pair_id/expected_rule_id/topic/…）→ 原逻辑
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip().lower() for c in rows[0]]
    if _is_official_answer_header(header):
        return _parse_official_answer_rows(rows)

    def idx(name: str) -> int | None:
        return header.index(name) if name in header else None

    def cell(row: tuple, name: str) -> Any:
        i = idx(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    def to_int(v: Any) -> int | None:
        try:
            return int(float(v)) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    expected: list[ExpectedDiff] = []
    for row in rows[1:]:
        if not any(cell(row, n) for n in ("pair_id", "expected_rule_id", "topic")):
            continue
        expected.append(
            ExpectedDiff(
                pair_id=str(cell(row, "pair_id") or "").strip(),
                company=str(cell(row, "company") or "").strip(),
                expected_rule_id=str(cell(row, "expected_rule_id") or "").strip(),
                topic=str(cell(row, "topic") or "").strip(),
                expected_severity=str(cell(row, "expected_severity") or "").strip().lower(),
                a_page=to_int(cell(row, "a_page")),
                h_page=to_int(cell(row, "h_page")),
                note=str(cell(row, "note") or "").strip(),
            )
        )
    return expected


_OFFICIAL_HEADER_KEYS = ("pdf页码", "原始数字", "错误数字")


def _is_official_answer_header(header: list[str]) -> bool:
    joined = "".join(header)
    return all(key in joined for key in _OFFICIAL_HEADER_KEYS)


def load_official_answer_key(path: Path) -> list[ExpectedDiff]:
    """读取主办方官方错误清单（列：序号/PDF页码/描述/原始数字/错误数字/差异额/变动说明）。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    return _parse_official_answer_rows(rows) if rows else []


def _parse_official_answer_rows(rows: list[tuple]) -> list[ExpectedDiff]:
    header = [str(c or "").strip().lower() for c in rows[0]]

    def idx_of(*names: str) -> int | None:
        for i, cell in enumerate(header):
            if any(name in cell for name in names):
                return i
        return None

    page_idx = idx_of("pdf页码", "页码")
    desc_idx = idx_of("描述")
    orig_idx = idx_of("原始数字", "原始")
    tamp_idx = idx_of("错误数字", "错误")
    note_idx = idx_of("变动说明", "说明")

    def cell(row: tuple, i: int | None) -> str:
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    expected: list[ExpectedDiff] = []
    for row in rows[1:]:
        page_text = cell(row, page_idx)
        tampered = cell(row, tamp_idx)
        if not page_text or not tampered:
            continue
        try:
            page = int(float(page_text))
        except ValueError:
            continue
        expected.append(
            ExpectedDiff(
                topic=cell(row, desc_idx),
                page=page,
                a_page=page,  # 植入错误都在 A 股侧，页码同时供旧的页码接近判断使用
                original_value=cell(row, orig_idx),
                tampered_value=tampered,
                description=cell(row, desc_idx),
                note=cell(row, note_idx),
            )
        )
    return expected


# ============================================================
# 数值 token 化
# ============================================================

def _canonical_number(text: str) -> str | None:
    """把一个数字文本规范化为可比较的标准形式。

    '126,411' / '126411' → '126411'；'(103,900)' → '103900'；'7.00' / '7.0' → '7'。
    括号负号/百分号在两侧都被剥离，因为答案清单与 diff 的符号约定并不一致，
    符号方向由页码 + 值本身共同约束已足够。
    """
    cleaned = str(text).replace(",", "").replace("(", "").replace(")", "").replace("%", "").strip()
    cleaned = cleaned.lstrip("+-")
    if not cleaned:
        return None
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if value == int(value):
        return str(int(value))
    return repr(value)


def _number_tokens(text: str | None) -> set[str]:
    """从自由文本里抽出所有数字并规范化成 token 集合。"""
    if not text:
        return set()
    tokens: set[str] = set()
    for raw in _NUMBER_TOKEN_RE.findall(str(text)):
        canonical = _canonical_number(raw)
        if canonical is not None:
            tokens.add(canonical)
    return tokens


def _diff_claimed_values(diff: Diff) -> set[str]:
    """diff **自己声称的**数值集合。

    只取结构化字段与系统生成的 summary —— 刻意**不含 evidence.snippet**：
    snippet 是 PDF 原文整行，含该行所有数字，用它做命中判定会让同页任意
    无关差异都算命中（这是旧实现召回虚高的主因）。
    """
    tokens: set[str] = set()
    for value in (diff.a_value, diff.h_value):
        if value is None:
            continue
        canonical = _canonical_number(str(value))
        if canonical is not None:
            tokens.add(canonical)
    for text in (diff.summary.zh, diff.summary.en):
        tokens |= _number_tokens(text)
    return tokens


# ============================================================
# 可见性
# ============================================================

def _norm_enum(value: Any) -> str:
    return str(getattr(value, "value", value) or "").strip().lower()


def is_visible_diff(diff: Diff) -> bool:
    """产品是否会把这条 diff 呈现为「需要处理的问题」。

    召回率必须按这个口径算：一条被降级为 expected/unresolved 或 severity=INFO/LOW
    的 diff，审计师在界面上看不到它，等同于漏检。
    """
    if _norm_enum(getattr(diff, "triage", "")) != "real":
        return False
    return _norm_enum(diff.severity) in _VISIBLE_SEVERITIES


# ============================================================
# 单条匹配
# ============================================================

def _match_by_values(diff: Diff, exp: ExpectedDiff) -> tuple[str, str]:
    """官方错误清单的值匹配：页码接近 + 数值 token 精确相等 + 方向正确。"""
    if not exp.tampered_value and not exp.original_value:
        return "", ""

    a_page, _ = _diff_pages(diff)
    # exp.page 是 A 股页码；只能与 diff 的 A 侧页码比。
    # 旧实现在 A 侧缺页码时拿 H 页码兜底 —— A/H 分页完全不同，该比较无意义。
    if exp.page is not None:
        if a_page is None:
            return "", ""
        if abs(a_page - exp.page) > _VALUE_PAGE_TOLERANCE:
            return "", ""

    claimed = _diff_claimed_values(diff)
    tampered = _canonical_number(exp.tampered_value) if exp.tampered_value else None
    original = _canonical_number(exp.original_value) if exp.original_value else None

    tampered_hit = tampered is not None and tampered in claimed
    original_hit = original is not None and original in claimed

    if tampered_hit and original_hit:
        return "exact", f"错误值 {exp.tampered_value} 与原始值 {exp.original_value} 均命中"
    if tampered_hit:
        return "value", f"错误值 {exp.tampered_value} 命中"
    if original_hit:
        # 只报出了正确值 —— 系统并没有指认篡改，不算发现。
        return "weak", f"仅命中原始值 {exp.original_value}（未指认错误值，不计入召回）"
    return "", ""


def _diff_pages(diff: Diff) -> tuple[int | None, int | None]:
    a_page = None
    h_page = None
    for ev in diff.evidence:
        if ev.side == ReportSide.A_SHARE and a_page is None:
            a_page = ev.page
        elif ev.side == ReportSide.H_SHARE and h_page is None:
            h_page = ev.page
    return a_page, h_page


def _topic_keys(text: str) -> set[str]:
    """提取文本中的 glossary canonical_key，用于跨语言主题模糊匹配。复用 glossary 术语表。"""
    if not text:
        return set()
    norm = to_simplified(text).lower()
    keys: set[str] = set()
    for form, canonical in glossary._to_canonical.items():
        if len(form) >= 3 and form in norm:
            keys.add(canonical)
    return keys


def _page_close(a: int | None, b: int | None, tol: int = _PAGE_TOLERANCE) -> bool:
    """一侧未指定页码时不作为否决条件（答案可能只标了一侧）。"""
    if a is None or b is None:
        return True
    return abs(a - b) <= tol


def _match_diff_to_expected(diff: Diff, exp: ExpectedDiff) -> tuple[str, str]:
    """返回 (匹配级别, 原因)；不匹配返回 ("", "")。"""
    # 0. 官方错误清单：按 页码 + 数值 token 匹配（优先级最高，命中即返回）
    level, reason = _match_by_values(diff, exp)
    if level:
        return level, reason

    a_page, h_page = _diff_pages(diff)
    pages_ok = _page_close(a_page, exp.a_page) and _page_close(h_page, exp.h_page)

    # 1. rule_id 精确
    if exp.expected_rule_id and diff.rule_id:
        if exp.expected_rule_id == diff.rule_id:
            return ("exact", f"rule_id 精确匹配 {diff.rule_id}") if pages_ok else ("", "rule_id 匹配但页码不接近")
        # 2. rule_id 前缀
        if exp.expected_rule_id in diff.rule_id or diff.rule_id in exp.expected_rule_id:
            return ("rule", f"rule_id 前缀匹配 {diff.rule_id}") if pages_ok else ("", "rule_id 前缀匹配但页码不接近")

    # 3. 模糊：主题 canonical_key 交集 + 页码。**不计入命中**，只转人工确认。
    exp_keys = _topic_keys(exp.topic)
    if exp_keys:
        diff_texts = " ".join(
            filter(None, [
                diff.topic.zh, diff.topic.en, diff.summary.zh, diff.summary.en,
                diff.diff_explanation.headline if diff.diff_explanation else "",
            ])
        )
        diff_keys = _topic_keys(diff_texts)
        overlap = exp_keys & diff_keys
        if overlap and pages_ok:
            return "fuzzy", f"主题关键词交集 {overlap}，页码接近（需人工确认）"
    return "", ""


# ============================================================
# 全局最优分配
# ============================================================

def _assign(
    expected: list[ExpectedDiff],
    diffs: list[Diff],
    levels: tuple[str, ...],
) -> dict[int, tuple[int, str, str]]:
    """在 expected × diffs 之间做全局最优一对一分配。

    只考虑 `levels` 中的匹配级别。返回 {expected_idx: (diff_idx, level, reason)}。

    旧实现是按 expected 顺序的贪心分配且不回溯：一条本可唯一匹配 #7 的 diff
    若被 #3 抢走，#7 就被记为漏检 —— 指标随答案行顺序波动、不可复现。
    """
    if not expected or not diffs:
        return {}

    scores: dict[tuple[int, int], tuple[float, str, str]] = {}
    for ei, exp in enumerate(expected):
        for di, diff in enumerate(diffs):
            level, reason = _match_diff_to_expected(diff, exp)
            if level in levels:
                scores[(ei, di)] = (float(_LEVEL_RANK[level]), level, reason)
    if not scores:
        return {}

    try:
        import numpy as np
        from scipy.optimize import linear_sum_assignment

        matrix = np.zeros((len(expected), len(diffs)), dtype=float)
        for (ei, di), (weight, _, _) in scores.items():
            matrix[ei, di] = weight
        rows, cols = linear_sum_assignment(matrix, maximize=True)
        assignment: dict[int, tuple[int, str, str]] = {}
        for ei, di in zip(rows, cols):
            entry = scores.get((int(ei), int(di)))
            if entry is None:
                continue  # 该格权重为 0，不是真实匹配
            _, level, reason = entry
            assignment[int(ei)] = (int(di), level, reason)
        return assignment
    except ImportError:
        # 无 scipy 时退回贪心（按权重降序全局取，仍优于旧的按行贪心）
        assignment = {}
        used_expected: set[int] = set()
        used_diffs: set[int] = set()
        for (ei, di), (weight, level, reason) in sorted(
            scores.items(), key=lambda kv: (-kv[1][0], kv[0])
        ):
            if ei in used_expected or di in used_diffs:
                continue
            used_expected.add(ei)
            used_diffs.add(di)
            assignment[ei] = (di, level, reason)
        return assignment


def evaluate(diffs: list[Diff], expected: list[ExpectedDiff], *, pair_id: str = "") -> EvalReport:
    """计算召回/精确率等指标。

    分三轮匹配，互不干扰：
    1. 严格命中：只在**用户可见**（triage=real 且 severity≥medium）的 diff 中做全局最优分配。
    2. 被压制的检出：剩余 expected 在**不可见** diff 中找匹配 → detected_but_suppressed。
    3. 转人工：剩余 expected 在全部 diff 中找 fuzzy/weak → 单列统计，不计入召回。
    """
    visible_idx = [i for i, d in enumerate(diffs) if is_visible_diff(d)]
    hidden_idx = [i for i, d in enumerate(diffs) if not is_visible_diff(d)]
    visible_diffs = [diffs[i] for i in visible_idx]
    hidden_diffs = [diffs[i] for i in hidden_idx]

    matches: list[MatchResult] = [
        MatchResult(expected=exp, match_level="missed", reason="未检出") for exp in expected
    ]
    used_diffs: set[int] = set()

    # --- 轮 1：严格命中（只看可见 diff）---
    strict = _assign(expected, visible_diffs, _STRICT_LEVELS)
    for ei, (local_di, level, reason) in strict.items():
        global_di = visible_idx[local_di]
        used_diffs.add(global_di)
        matches[ei] = MatchResult(
            expected=expected[ei],
            matched_diff=diffs[global_di],
            match_level=level,
            reason=reason,
        )

    # --- 轮 2：查到了但被自己压掉（FN 主战场）---
    pending = [i for i, m in enumerate(matches) if m.matched_diff is None]
    if pending and hidden_diffs:
        suppressed = _assign([expected[i] for i in pending], hidden_diffs, _STRICT_LEVELS)
        for local_ei, (local_di, level, reason) in suppressed.items():
            ei = pending[local_ei]
            global_di = hidden_idx[local_di]
            used_diffs.add(global_di)
            matches[ei].suppressed_diff = diffs[global_di]
            matches[ei].suppressed_level = level
            matches[ei].reason = (
                f"检出但被压制（triage={_norm_enum(diffs[global_di].triage)}, "
                f"severity={_norm_enum(diffs[global_di].severity)}）：{reason}"
            )

    # --- 轮 3：fuzzy / weak，转人工确认，不计入召回 ---
    pending = [
        i for i, m in enumerate(matches)
        if m.matched_diff is None and m.suppressed_diff is None
    ]
    remaining_idx = [i for i in range(len(diffs)) if i not in used_diffs]
    remaining = [diffs[i] for i in remaining_idx]
    if pending and remaining:
        soft = _assign([expected[i] for i in pending], remaining, ("fuzzy", "weak"))
        for local_ei, (local_di, level, reason) in soft.items():
            ei = pending[local_ei]
            global_di = remaining_idx[local_di]
            used_diffs.add(global_di)
            matches[ei].matched_diff = None  # 明确：不算命中
            matches[ei].match_level = level
            matches[ei].reason = reason
            matches[ei].suppressed_diff = diffs[global_di]

    hit_count = sum(1 for m in matches if m.match_level in _STRICT_LEVELS and m.matched_diff is not None)
    detected_but_suppressed = sum(
        1 for m in matches if m.matched_diff is None and m.suppressed_level in _STRICT_LEVELS
    )
    needs_manual_review = sum(1 for m in matches if m.match_level == "fuzzy")
    weak_count = sum(1 for m in matches if m.match_level == "weak")

    # --- 误报分档：hard / soft / suppressed 三档互不混淆 ---
    unmatched = [d for i, d in enumerate(diffs) if i not in used_diffs]
    hard_fp = [d for d in unmatched if _norm_enum(d.triage) == "real"]
    soft_fp = [d for d in unmatched if _norm_enum(d.triage) == "unresolved"]
    suppressed_fp = [d for d in unmatched if _norm_enum(d.triage) == "expected"]

    recall = hit_count / len(expected) if expected else 0.0
    precision = hit_count / (hit_count + len(hard_fp)) if (hit_count + len(hard_fp)) else 0.0

    return EvalReport(
        pair_id=pair_id,
        expected_count=len(expected),
        detected_count=len(diffs),
        hit_count=hit_count,
        false_positive_count=len(hard_fp),
        recall=round(recall, 4),
        precision=round(precision, 4),
        matches=matches,
        unmatched_diffs=hard_fp,
        hard_fp_count=len(hard_fp),
        soft_fp_count=len(soft_fp),
        suppressed_count=len(suppressed_fp),
        needs_manual_review_count=needs_manual_review,
        weak_count=weak_count,
        detected_but_suppressed_count=detected_but_suppressed,
        visible_diff_count=len(visible_diffs),
        recall_by_rule_id=_recall_by_rule_id(matches),
    )


def _recall_by_rule_id(matches: list[MatchResult]) -> dict[str, tuple[int, int]]:
    """按命中 diff 的 rule_id 统计命中分布。

    用来暴露"单一路径的 100% 掩盖其余路径的 0%" —— 旧基线正是这个问题。
    返回 {rule_id: (命中数, 该 rule 参与的匹配数)}。未命中的 expected 归入 "__missed__"。
    """
    stats: dict[str, list[int]] = {}
    for m in matches:
        if m.matched_diff is not None:
            key = m.matched_diff.rule_id or "__unlabeled__"
        elif m.suppressed_diff is not None:
            key = m.suppressed_diff.rule_id or "__unlabeled__"
        else:
            key = "__missed__"
        bucket = stats.setdefault(key, [0, 0])
        bucket[1] += 1
        if m.matched_diff is not None:
            bucket[0] += 1
    return {k: (v[0], v[1]) for k, v in stats.items()}


# ============================================================
# 导出与打印
# ============================================================

def export_eval_excel(report: EvalReport, out_path: Path) -> None:
    """导出评估明细 Excel：指标汇总 / 命中明细 / 漏检清单 / 误报清单。"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "指标汇总"
    ws.append(["指标", "数值", "说明"])
    miss_rate = round(1 - report.recall, 4) if report.expected_count else 0.0
    for k, v, note in [
        ("样本对", report.pair_id, ""),
        ("预期差异数", report.expected_count, ""),
        ("检出差异数", report.detected_count, "全部 triage"),
        ("其中用户可见", report.visible_diff_count, "triage=real 且 severity>=medium"),
        ("严格命中数", report.hit_count, "只认可见 diff"),
        ("检出但被压制", report.detected_but_suppressed_count, "查到了却被自己降级/压掉，FN 主战场"),
        ("待人工确认(fuzzy)", report.needs_manual_review_count, "不计入召回"),
        ("仅命中原始值(weak)", report.weak_count, "未指认错误值，不计入召回"),
        ("hard 误报", report.hard_fp_count, "未命中且 triage=real"),
        ("soft 误报", report.soft_fp_count, "未命中且 triage=unresolved"),
        ("自称可解释(suppressed)", report.suppressed_count, "未命中且 triage=expected，需抽检确认抑制是否正确"),
        ("召回率", report.recall, ""),
        ("精确率", report.precision, "分母只含 hard 误报"),
        ("漏检率", miss_rate, ""),
    ]:
        ws.append([k, v, note])

    ws_rule = wb.create_sheet("按规则召回")
    ws_rule.append(["rule_id", "命中数", "参与匹配数"])
    for rule_id, (hits, total) in sorted(report.recall_by_rule_id.items()):
        ws_rule.append([rule_id, hits, total])

    ws_hit = wb.create_sheet("命中明细")
    ws_hit.append(["预期规则", "预期主题", "匹配级别", "命中差异ID", "命中rule_id", "匹配原因"])
    for m in report.matches:
        if m.matched_diff is not None:
            ws_hit.append([
                m.expected.expected_rule_id, m.expected.topic, m.match_level,
                m.matched_diff.diff_id, m.matched_diff.rule_id, m.reason,
            ])

    ws_sup = wb.create_sheet("检出但被压制")
    ws_sup.append(["预期主题", "匹配级别", "差异ID", "rule_id", "triage", "severity", "原因"])
    for m in report.matches:
        if m.matched_diff is None and m.suppressed_diff is not None:
            d = m.suppressed_diff
            ws_sup.append([
                m.expected.topic, m.suppressed_level or m.match_level, d.diff_id, d.rule_id,
                _norm_enum(d.triage), _norm_enum(d.severity), m.reason,
            ])

    ws_miss = wb.create_sheet("漏检清单")
    ws_miss.append(["预期规则", "预期主题", "预期严重度", "预期A页", "预期H页", "备注", "待补规则建议"])
    for m in report.matches:
        if m.matched_diff is not None:
            continue
        if m.expected.expected_rule_id:
            suggest = f"规则 {m.expected.expected_rule_id} 未触发，请检查规则配置与对齐"
        else:
            suggest = f"建议新增规则覆盖：{m.expected.topic}"
        ws_miss.append([
            m.expected.expected_rule_id, m.expected.topic, m.expected.expected_severity,
            m.expected.a_page, m.expected.h_page, m.expected.note, suggest,
        ])

    ws_fp = wb.create_sheet("误报清单")
    ws_fp.append(["差异ID", "rule_id", "严重度", "主题", "差异说明", "定位"])
    for d in report.unmatched_diffs:
        a_page, h_page = _diff_pages(d)
        ws_fp.append([d.diff_id, d.rule_id, d.severity.value, d.topic.best(), d.summary.best(), f"A{a_page}/H{h_page}"])

    wb.save(out_path)


def print_report(report: EvalReport) -> None:
    miss_rate = round(1 - report.recall, 4) if report.expected_count else 0.0
    print(
        f"[{report.pair_id or '-'}] 预期 {report.expected_count} / 检出 {report.detected_count}"
        f"（可见 {report.visible_diff_count}） / 命中 {report.hit_count}"
    )
    print(
        f"  召回率 {report.recall * 100:.1f}%  精确率 {report.precision * 100:.1f}%  "
        f"漏检率 {miss_rate * 100:.1f}%"
    )
    print(
        f"  误报分档：hard {report.hard_fp_count} / soft {report.soft_fp_count} / "
        f"自称可解释 {report.suppressed_count}"
    )
    if report.detected_but_suppressed_count:
        print(
            f"  [!] 检出但被自己压制 {report.detected_but_suppressed_count} 条 —— "
            f"这些是查到了却不会展示给用户的真实错误"
        )
    if report.needs_manual_review_count or report.weak_count:
        print(
            f"  待人工确认 fuzzy {report.needs_manual_review_count} / "
            f"仅命中原始值 weak {report.weak_count}（均不计入召回）"
        )
    missed = [m for m in report.matches if m.matched_diff is None and m.suppressed_diff is None]
    if missed:
        print(f"  完全漏检 {len(missed)} 条：")
        for m in missed:
            print(f"    - [{m.expected.expected_rule_id or '?'}] {m.expected.topic}")
