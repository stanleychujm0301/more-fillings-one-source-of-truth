from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from ahcc.report.excel import export_excel
from ahcc.report.pdf import export_pdf
from ahcc.report import _style as S
from ahcc.schemas import (
    Diff,
    DiffExplanation,
    DiffExplanationItem,
    DiffSeverity,
    DiffType,
    DisclosureCoverageItem,
    Evidence,
    Job,
    LocalizedString,
    ReportSide,
)


def _relative_luminance(hex_color: str) -> float:
    """Return WCAG relative luminance for a 6-digit RGB hex color."""
    raw = hex_color[-6:]
    channels = [int(raw[i:i + 2], 16) / 255 for i in (0, 2, 4)]

    def linear(value: float) -> float:
        return value / 12.92 if value <= 0.03928 else ((value + 0.055) / 1.055) ** 2.4

    r, g, b = [linear(value) for value in channels]
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def test_report_visual_palette_keeps_homepages_light() -> None:
    large_area_tokens = {
        "HERO_NAVY": S.HERO_NAVY,
        "HERO_BLUE": S.HERO_BLUE,
        "HERO_CYAN": S.HERO_CYAN,
        "HERO_STEEL": S.HERO_STEEL,
        "HERO_MIST": S.HERO_MIST,
        "HERO_WASH": S.HERO_WASH,
        "REPORT_SURFACE": S.REPORT_SURFACE,
        "REPORT_PANEL": S.REPORT_PANEL,
        "REPORT_PANEL_SOFT": S.REPORT_PANEL_SOFT,
        "REPORT_GLOW": S.REPORT_GLOW,
        "REPORT_MINT": S.REPORT_MINT,
        "REPORT_WARM": S.REPORT_WARM,
    }
    dark_surfaces = {
        name: color
        for name, color in large_area_tokens.items()
        if _relative_luminance(color) < 0.72
    }

    assert dark_surfaces == {}
    assert S.MONO_RAMP[0] != S.KPMG_BLUE
    assert min(_relative_luminance(color) for color in S.MONO_RAMP) >= 0.42


def test_report_uses_light_editorial_rules_instead_of_heavy_blue_bars() -> None:
    pdf_source = Path("ahcc/report/pdf.py").read_text(encoding="utf-8")

    assert S.REPORT_SECTION_FILL != S.KPMG_BLUE
    assert S.REPORT_SECTION_RULE != S.KPMG_BLUE
    assert _relative_luminance(S.REPORT_SECTION_FILL) > 0.9
    assert _relative_luminance(S.REPORT_SECTION_RULE) > 0.45
    assert '("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + S.KPMG_BLUE))' not in pdf_source
    assert "S.REPORT_SECTION_FILL" in pdf_source
    assert "S.REPORT_SECTION_RULE" in pdf_source


def test_pdf_cover_uses_editorial_grid_not_decorative_orbs() -> None:
    pdf_source = Path("ahcc/report/pdf.py").read_text(encoding="utf-8")

    assert "canvas.circle(" not in pdf_source
    assert "COVER_GRID_LINE" in pdf_source
    assert "COVER_PANEL_TINT" in pdf_source


def test_format_duration_renders_human_readable():
    """核查耗时格式化：None/0 → —，<60s → 秒，>=60s → 分秒。"""
    assert S.format_duration(None) == "—"
    assert S.format_duration(0) == "—"
    assert S.format_duration(-3) == "—"
    assert S.format_duration(12.34) == "12.3 秒"
    assert S.format_duration(59.9) == "59.9 秒"
    assert S.format_duration(60) == "1 分 0 秒"
    assert S.format_duration(222.5) == "3 分 42 秒"


def test_report_temp_dir_helper_creates_writable_directory() -> None:
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    tmp_dir = S.make_report_temp_dir(out_dir, "chart")

    try:
        probe = tmp_dir / "probe.txt"
        probe.write_text("ok", encoding="utf-8")
        assert probe.read_text(encoding="utf-8") == "ok"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _explained_diff() -> Diff:
    return Diff(
        diff_id="d-explained",
        diff_type=DiffType.DISCLOSURE,
        severity=DiffSeverity.MEDIUM,
        triage="real",
        topic=LocalizedString(zh="利润分配", en="Profit distribution"),
        summary=LocalizedString(zh="事件事实不一致：利润分配", en="Event mismatch"),
        diff_explanation=DiffExplanation(
            headline="利润分配股利总额不一致",
            issue="A 披露共计股利人民币2,503,994千元；H 披露 RMB25,039,945 thousand",
            location="A 第453页；H 第453页",
            items=[
                DiffExplanationItem(
                    label="股利总额",
                    role="dividend_total",
                    a_value=2_503_994_000,
                    h_value=25_039_945_000,
                    delta=22_535_951_000,
                    a_page=453,
                    h_page=453,
                    a_snippet="共计股利人民币2,503,994千元",
                    h_snippet="total dividends amounting to RMB25,039,945 thousand",
                )
            ],
            review_hint="优先核对利润分配附注中的股利总额。",
        ),
        evidence=[
            Evidence(side=ReportSide.A_SHARE, page=453, snippet="共计股利人民币2,503,994千元"),
            Evidence(side=ReportSide.H_SHARE, page=453, snippet="total dividends amounting to RMB25,039,945 thousand"),
        ],
        rule_id="event_fact_match",
    )


def _report_job() -> Job:
    return Job(
        job_id="j-premium-report",
        company_name="示例金融项目",
        a_file="a.pdf",
        h_file="h.pdf",
        status="done",
        diffs=[
            _explained_diff().model_copy(update={"diff_id": "D-001", "severity": DiffSeverity.HIGH}),
            *_sample_diffs(),
        ],
        coverage_items=[
            DisclosureCoverageItem(
                coverage_id="c-premium",
                category="metric",
                status="matched",
                topic=LocalizedString(zh="员工人数"),
                note="A/H 披露口径匹配。",
            )
        ],
        comparison_summary={
            "total_diff_count": 5,
            "real_diff_count": 3,
            "expected_diff_count": 1,
            "unresolved_diff_count": 1,
            "coverage_count": 1,
            "warning_count": 0,
        },
        started_at=datetime(2026, 6, 25, 9, 0, tzinfo=timezone.utc),
        finished_at=datetime(2026, 6, 25, 9, 3, 42, tzinfo=timezone.utc),
        duration_seconds=222.5,
    )


def test_excel_export_strips_pdf_control_characters() -> None:
    dirty_text = "中国平安\x01综合金融模式\x02覆盖330个网点"
    job = Job(
        job_id="j-dirty",
        a_file="a.pdf",
        h_file="h.pdf",
        diffs=[
            Diff(
                diff_id="d-1",
                diff_type=DiffType.DISCLOSURE,
                severity=DiffSeverity.MEDIUM,
                triage="real",
                topic=LocalizedString(zh="经营情况"),
                summary=LocalizedString(zh=dirty_text),
                evidence=[
                    Evidence(side=ReportSide.A_SHARE, page=13, snippet=dirty_text),
                    Evidence(side=ReportSide.H_SHARE, page=15, snippet=dirty_text),
                ],
            )
        ],
        coverage_items=[
            DisclosureCoverageItem(
                coverage_id="c-1",
                category="event",
                status="matched",
                topic=LocalizedString(zh="渠道覆盖"),
                note=dirty_text,
            )
        ],
        profile_a={
            "doc_id": "A",
            "total_pages": 10,
            "metrics": [
                {
                    "canonical_key": "branch_count",
                    "name": {"zh": "网点数量"},
                    "value": 330,
                    "page": 13,
                    "evidence": {"snippet": dirty_text},
                }
            ],
        },
        profile_h={
            "doc_id": "H",
            "total_pages": 10,
            "narratives": [
                {
                    "topic_key": "operations",
                    "topic_label": "经营情况",
                    "word_count": 20,
                    "page_range": [15, 15],
                    "summary": dirty_text,
                }
            ],
        },
        comparison_summary={
            "warnings": [{"side": "H", "flag": "dirty_text", "message": dirty_text}],
        },
    )
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)

        workbook = load_workbook(out_path, read_only=True)
        try:
            illegal_chars = {"\x01", "\x02"}
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if isinstance(value, str):
                            assert not any(char in value for char in illegal_chars)
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_excel_export_writes_diff_explanation_columns() -> None:
    job = Job(job_id="j-explained", a_file="a.pdf", h_file="h.pdf", diffs=[_explained_diff()])
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)

        workbook = load_workbook(out_path, read_only=True)
        try:
            ws = workbook.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
            assert "差异说明" in headers
            assert "A定位与取值" in headers
            assert "H定位与取值" in headers
            assert "审阅提示" in headers
            assert any("利润分配股利总额不一致" in str(value) for value in row)
            assert any("A 第453页" in str(value) and "2,503,994" in str(value) for value in row)
            assert any("H 第453页" in str(value) and "25,039,945" in str(value) for value in row)
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_excel_export_uses_bilingual_side_headers() -> None:
    job = Job(
        job_id="j-bilingual-export",
        check_mode="h_bilingual",
        a_file="zh.pdf",
        h_file="en.pdf",
        diffs=[_explained_diff()],
    )
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)

        workbook = load_workbook(out_path, read_only=True)
        try:
            ws = workbook["差异清单"]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            assert "中文定位与取值" in headers
            assert "英文定位与取值" in headers
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_pdf_export_accepts_diff_explanation() -> None:
    job = Job(job_id="j-pdf-explained", a_file="a.pdf", h_file="h.pdf", diffs=[_explained_diff()])
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-{uuid4().hex}.pdf"

    try:
        export_pdf(job, out_path)

        assert out_path.exists()
        assert out_path.stat().st_size > 0
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_pdf_export_uses_executive_report_structure() -> None:
    fitz = pytest.importorskip("fitz")
    job = _report_job()
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-premium-{uuid4().hex}.pdf"

    try:
        export_pdf(job, out_path)

        with fitz.open(out_path) as doc:
            text = "\n".join(page.get_text() for page in doc)
            assert doc.page_count >= 2
            assert doc.page_count <= 3
            page_dark_ratios = []
            for page in doc:
                pix = page.get_pixmap(matrix=fitz.Matrix(0.45, 0.45), alpha=False)
                pixels = pix.samples
                dark_pixels = sum(
                    1
                    for i in range(0, len(pixels), 3)
                    if pixels[i] < 24 and pixels[i + 1] < 42 and pixels[i + 2] < 82
                )
                page_dark_ratios.append(dark_pixels / (pix.width * pix.height))
            cover = doc[0].get_pixmap(matrix=fitz.Matrix(0.45, 0.45), alpha=False)
            pixels = cover.samples
            light_pixels = sum(
                1
                for i in range(0, len(pixels), 3)
                if pixels[i] > 238 and pixels[i + 1] > 242 and pixels[i + 2] > 248
            )
            assert page_dark_ratios[0] < 0.04
            assert max(page_dark_ratios) < 0.06
            assert light_pixels / (cover.width * cover.height) > 0.55
        assert "项目名称" in text
        assert "EXECUTIVE REPORT" in text
        assert "ONE PAGE REVIEW" in text
        assert "SIGNAL STRIP" in text
        assert "REVIEW FLOW" in text
        assert "EXECUTIVE PULSE" in text
        assert "执行摘要" in text
        assert "审阅动线" in text
        assert "证据定位" in text
        assert "差异总览" in text
        assert "审阅提示" in text or "证据链" in text
        assert "A 侧证据" in text
        assert "H 侧证据" in text
        assert "复核动作" in text
        assert "北京时间" in text
        assert "D-00\n1" not in text
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def _sample_diffs() -> list[Diff]:
    """构造覆盖各级别严重度与分流的差异，用于视觉冒烟测试。"""
    return [
        Diff(
            diff_id="d-critical",
            diff_type=DiffType.NUMERIC,
            severity=DiffSeverity.CRITICAL,
            triage="real",
            topic=LocalizedString(zh="营业收入"),
            summary=LocalizedString(zh="A/H 营业收入存在重大差异"),
            a_value=100_000_000.0,
            h_value=98_500_000.0,
            delta=-1_500_000.0,
            evidence=[Evidence(side=ReportSide.A_SHARE, page=12, snippet="营收 1 亿")],
        ),
        Diff(
            diff_id="d-high",
            diff_type=DiffType.CROSS_CHECK,
            severity=DiffSeverity.HIGH,
            triage="real",
            topic=LocalizedString(zh="总资产勾稽"),
            summary=LocalizedString(zh="总资产不等于负债加权益"),
            evidence=[Evidence(side=ReportSide.A_SHARE, page=34, snippet="总资产 50 亿")],
        ),
        Diff(
            diff_id="d-medium",
            diff_type=DiffType.DISCLOSURE,
            severity=DiffSeverity.MEDIUM,
            triage="expected",
            topic=LocalizedString(zh="研发费用资本化"),
            summary=LocalizedString(zh="披露口径差异"),
            evidence=[Evidence(side=ReportSide.H_SHARE, page=56, snippet="R&D")],
        ),
        Diff(
            diff_id="d-low",
            diff_type=DiffType.STANDARD,
            severity=DiffSeverity.LOW,
            triage="unresolved",
            topic=LocalizedString(zh="折旧年限"),
            summary=LocalizedString(zh="年限披露差异"),
            evidence=[Evidence(side=ReportSide.A_SHARE, page=78, snippet="5 年")],
        ),
    ]


def test_pdf_excel_smoke_generates_files() -> None:
    """导出 PDF/Excel 样例，验证文件非空且可加载。"""
    job = Job(
        job_id="j-smoke",
        company_name="示例股份",
        a_file="a.pdf",
        h_file="h.pdf",
        diffs=_sample_diffs(),
        coverage_items=[
            DisclosureCoverageItem(
                coverage_id="c-1",
                category="metric",
                status="matched",
                topic=LocalizedString(zh="员工人数"),
                note="匹配",
            )
        ],
        comparison_summary={
            "total_diff_count": 4,
            "real_diff_count": 2,
            "expected_diff_count": 1,
            "unresolved_diff_count": 1,
            "coverage_count": 1,
            "warning_count": 0,
        },
    )
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / f"report-smoke-{uuid4().hex}.pdf"
    xlsx_path = out_dir / f"report-smoke-{uuid4().hex}.xlsx"

    try:
        export_pdf(job, pdf_path)
        export_excel(job, xlsx_path)
        assert pdf_path.exists() and pdf_path.stat().st_size > 0
        assert xlsx_path.exists() and xlsx_path.stat().st_size > 0

        workbook = load_workbook(xlsx_path, read_only=True)
        try:
            assert "核查总览" in workbook.sheetnames
            assert "差异清单" in workbook.sheetnames
        finally:
            workbook.close()
    finally:
        for p in (pdf_path, xlsx_path):
            if p.exists():
                try:
                    p.unlink()
                except PermissionError:
                    pass


def test_excel_conditional_formatting_applied() -> None:
    """检查「差异」列是否应用了数据条条件格式。"""
    job = Job(job_id="j-databar", a_file="a.pdf", h_file="h.pdf", diffs=_sample_diffs())
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-databar-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)
        workbook = load_workbook(out_path)
        try:
            ws = workbook["差异清单"]
            delta_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "差异":
                    delta_col = idx
                    break
            assert delta_col is not None
            assert len(ws.conditional_formatting) > 0
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_excel_overview_uses_dashboard_layout_and_beijing_metadata() -> None:
    job = _report_job()
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-dashboard-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)
        workbook = load_workbook(out_path)
        try:
            ws = workbook["核查总览"]
            values = [
                str(cell.value)
                for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=9)
                for cell in row
                if cell.value is not None
            ]
            joined = "\n".join(values)
            assert ws.sheet_view.showGridLines is False
            assert ws.sheet_view.zoomScale == 90
            assert "项目名称" in joined
            assert "示例金融项目" in joined
            assert "核查模式" in joined
            assert "北京时间" in joined
            assert "执行摘要" in joined
            assert "ONE PAGE REVIEW" in joined
            assert "Evidence-led disclosure consistency review" in joined
            assert "审阅提示" in joined
            assert "审阅动线" in joined
            assert "证据定位" in joined
            assert ws.freeze_panes == "A2"
            assert ws["A1"].fill.fgColor.rgb.endswith(S.REPORT_SURFACE)
            assert ws["A1"].font.color.rgb.endswith(S.INK)
            assert ws["A2"].fill.fgColor.rgb.endswith(S.HERO_WASH)
            assert ws["G2"].value == "ONE PAGE REVIEW"
            assert ws["G3"].value == "需重点复核"
            assert ws["G3"].fill.fgColor.rgb.endswith(S.REPORT_PANEL)
            assert ws["G3"].border.left.color.rgb.endswith(S.REPORT_PANEL_BORDER)
            assert ws["A9"].value == "项目名称"
            assert ws["C9"].value == "核查模式"
            assert ws["E9"].value == "生成时间"
            assert ws["G9"].value == "核查耗时"
            chart_anchor_row = None
            distribution_header_row = None
            for candidate in range(1, ws.max_row + 1):
                if ws.cell(row=candidate, column=1).value == "分布概览":
                    distribution_header_row = candidate
                    chart_anchor_row = candidate + 1
                    break
            assert chart_anchor_row is not None
            assert distribution_header_row is not None
            assert ws.cell(row=distribution_header_row, column=1).fill.fgColor.rgb.endswith(S.REPORT_SECTION_FILL)
            assert ws.cell(row=distribution_header_row, column=1).border.bottom.color.rgb.endswith(S.REPORT_SECTION_RULE)
            assert ws.cell(row=chart_anchor_row, column=1).fill.fgColor.rgb.endswith(S.REPORT_PANEL)
            assert ws.cell(row=chart_anchor_row, column=6).fill.fgColor.rgb.endswith(S.REPORT_PANEL)
            dark_fills = []
            for row in ws.iter_rows(min_row=1, max_row=26, min_col=1, max_col=9):
                for cell in row:
                    rgb = cell.fill.fgColor.rgb or ""
                    rgb = rgb[-6:] if rgb and rgb != "00000000" else "FFFFFF"
                    red = int(rgb[0:2], 16)
                    green = int(rgb[2:4], 16)
                    blue = int(rgb[4:6], 16)
                    if red < 55 and green < 70 and blue < 100:
                        dark_fills.append((cell.coordinate, rgb))
            assert dark_fills == []
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_excel_overview_embeds_rendered_distribution_charts() -> None:
    job = _report_job()
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-chart-images-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)
        workbook = load_workbook(out_path)
        try:
            ws = workbook["核查总览"]
            assert len(ws._images) >= 1
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass


def test_excel_tables_use_polished_table_contract() -> None:
    job = _report_job()
    out_dir = Path("storage") / "test-artifacts"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"report-table-style-{uuid4().hex}.xlsx"

    try:
        export_excel(job, out_path)
        workbook = load_workbook(out_path)
        try:
            ws = workbook["差异清单"]
            headers = [cell.value for cell in ws[1]]
            assert headers[:4] == ["差异ID", "严重度", "分流", "差异类型"]
            assert ws.freeze_panes == "A2"
            assert ws.auto_filter.ref == ws.dimensions
            assert ws.sheet_view.showGridLines is False
            assert ws.sheet_view.zoomScale == 85
            assert ws.sheet_properties.tabColor is not None
            assert ws.row_dimensions[1].height >= 34
            assert ws.row_dimensions[2].height >= 36
            assert ws["A1"].font.size >= 12
            assert ws["A1"].font.color.rgb.endswith(S.INK)
            assert ws["A1"].fill.fgColor.rgb.endswith(S.REPORT_PANEL)
            assert ws["A1"].border.bottom.style == "medium"
            assert ws["A1"].border.bottom.color.rgb.endswith(S.HEADER_BOTTOM)
            assert ws["B2"].fill.fgColor.rgb.endswith(S.REPORT_PANEL_SOFT)
            assert ws["C2"].fill.fgColor.rgb.endswith(S.REPORT_PANEL_SOFT)
            assert ws["A2"].border.left.style in {"thin", "medium"}
        finally:
            workbook.close()
    finally:
        if out_path.exists():
            try:
                out_path.unlink()
            except PermissionError:
                pass
