"""真实样本慢速回归 —— 项目的两条核心验收线。

运行方式（默认 pytest 跳过；需要本机有样本 PDF）：

    AHCC_SAMPLES_DIR="F:/毕马威黑客松/样本测试" python -m pytest -m slow -q

覆盖：
1. 主办方 3 组含错误样本（光大银行/长城汽车/青岛啤酒，各 15 处植入错误）：
   文本层叠加篡改检测应 45/45 全检出、页码与数值对与官方错误清单一致、零多报；
2. 干净年报 PDF 上叠加检测零误报；
3. 光大银行真实 A+H 年报：分支机构核查应稳定产出恰好 40 条 BRANCH_* 差异
   （H 稿「资产规模」整列被打乱，详见对应用例的 docstring）。
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest

pytestmark = pytest.mark.slow

_SAMPLES_DIR = os.environ.get("AHCC_SAMPLES_DIR", "")


def _samples_root() -> Path:
    if not _SAMPLES_DIR:
        pytest.skip("AHCC_SAMPLES_DIR not set")
    root = Path(_SAMPLES_DIR)
    if not root.is_dir():
        pytest.skip(f"AHCC_SAMPLES_DIR not found: {root}")
    return root


def _normalize(text: str) -> str:
    return str(text).replace(",", "").replace("(", "").replace(")", "").replace("%", "").strip()


def _load_expected(xlsx: Path) -> list[tuple[int, str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    expected = []
    for row in list(wb.active.iter_rows(values_only=True))[1:]:
        if row[0] is None:
            continue
        expected.append((int(row[1]), _normalize(str(row[3])), _normalize(str(row[4]))))
    return expected


_SAMPLE_SETS = [
    ("光大银行2025年A股年报_含错误_测试版.pdf", "光大银行2025年A股年报_错误清单_15处.xlsx"),
    ("长城汽车A股年报_含错误_测试版.pdf", "长城汽车A股年报_错误清单_15处.xlsx"),
    ("青岛啤酒A股年报_含错误_测试版.pdf", "青岛啤酒A股年报_错误清单_15处.xlsx"),
]


@pytest.mark.parametrize("pdf_name,xlsx_name", _SAMPLE_SETS)
def test_overlay_detects_all_planted_errors_with_zero_extras(pdf_name: str, xlsx_name: str) -> None:
    from ahcc.check.text_overlay_tamper import scan_pdf_overlays

    root = _samples_root()
    pdf = root / "sample" / pdf_name
    xlsx = root / "sample" / xlsx_name
    if not pdf.is_file() or not xlsx.is_file():
        pytest.skip(f"sample files missing: {pdf_name}")

    expected = _load_expected(xlsx)
    hits = scan_pdf_overlays(str(pdf))

    matched = 0
    used: set[int] = set()
    for page, orig, tamp in expected:
        for idx, hit in enumerate(hits):
            if idx in used:
                continue
            pair = {_normalize(hit.visible_value), _normalize(hit.hidden_value)}
            if hit.page == page and pair == {orig, tamp}:
                # 叠加层判定应指认"错误数字"为可见值
                assert _normalize(hit.visible_value) == tamp, (
                    f"p{page}: visible={hit.visible_value} expected tampered={tamp}"
                )
                matched += 1
                used.add(idx)
                break

    assert matched == len(expected) == 15, f"检出 {matched}/{len(expected)}"
    assert len(hits) == 15, f"多报 {len(hits) - 15} 条"


_CLEAN_PDFS = [
    "sample/光大银行_2025年H股年报.pdf",
    "sample/长城汽车_2024年H股年报.pdf",
    "sample/青岛啤酒_2024年H股年报.pdf",
    "光大银行/A 中国光大银行股份有限公司2025年年度报告.pdf",
    "光大银行/H 中国光大银行股份有限公司2025年年度报告 2.pdf",
    "中国平安/A 中国平安2025年年度报告.pdf",
]


@pytest.mark.parametrize("rel_path", _CLEAN_PDFS)
def test_overlay_zero_false_positives_on_clean_reports(rel_path: str) -> None:
    from ahcc.check.text_overlay_tamper import scan_pdf_overlays

    pdf = _samples_root() / rel_path
    if not pdf.is_file():
        pytest.skip(f"clean pdf missing: {rel_path}")

    hits = scan_pdf_overlays(str(pdf))

    assert hits == [], f"误报 {len(hits)} 条: {[(h.page, h.visible_value, h.hidden_value) for h in hits[:5]]}"


def test_branch_checks_find_exactly_40_diffs_on_ceb_real_pair() -> None:
    """光大银行真实 A+H 年报 —— 用户口径的「本身已有错误」：分支机构 40 处不一致。

    这 40 处是真差异，不是抽取错位。判据（本用例一并固化）：

    - 44/44 行的机构数量两侧一致；
    - 44/44 行的办公地址两侧一致；
    - 跨名称命中 40 行，其中锚点跟着数值一起移动的 0 行；
    - 两侧资产规模多重集完全相等（纯排列）。

    另有外部佐证：官方正式版 H 报告（316 页 InDesign 版）与 A 报告逐行一致，
    说明被核查的这份 H 稿（352 页 Word 版）「资产规模」整列被打乱。
    """
    from ahcc.check.branch_disclosure import (
        _branch_row_alignment,
        extract_branch_table,
        load_branch_lightweight_doc,
        run_branch_checks,
    )
    from ahcc.schemas import ReportSide

    root = _samples_root()
    a_pdf = root / "光大银行" / "A 中国光大银行股份有限公司2025年年度报告.pdf"
    h_pdf = root / "光大银行" / "H 中国光大银行股份有限公司2025年年度报告 2.pdf"
    if not a_pdf.is_file() or not h_pdf.is_file():
        pytest.skip("CEB real A/H pair missing")

    diffs, diagnostics = run_branch_checks(str(a_pdf), str(h_pdf))

    assert len(diffs) == 40, f"expected 40 branch diffs, got {len(diffs)}"
    assert all(d.rule_id == "branch_asset_scale_match" for d in diffs)
    assert all(d.triage == "real" for d in diffs)
    assert diagnostics["branch_diff_count"] == 40
    assert "branch_alignment_warning" not in diagnostics, "真实差异不应被记成抽取预警"

    # 判据本身：锚点稳定、多重集相等 —— 这才是「列错乱」而非「行错位」的证据
    a_branches = extract_branch_table(load_branch_lightweight_doc(str(a_pdf), ReportSide.A_SHARE))
    h_branches = extract_branch_table(load_branch_lightweight_doc(str(h_pdf), ReportSide.H_SHARE))
    matched = sorted(set(a_branches) & set(h_branches))
    assert len(matched) == 44
    assert all(a_branches[n]["count"] == h_branches[n]["count"] for n in matched)
    assert all(a_branches[n]["address"] == h_branches[n]["address"] for n in matched)

    alignment = _branch_row_alignment(a_branches, h_branches)
    assert alignment.anchors_available
    assert alignment.anchor_moved_hits == []
    assert len(alignment.anchor_stable_hits) == 40
    assert alignment.value_multiset_equal
    assert alignment.is_column_permuted
    assert alignment.comparable


def test_branch_checks_find_40_diffs_on_sample_pair() -> None:
    """sample 目录里的光大银行对用的是同一份 H 稿，因此同样是 40 处。"""
    from ahcc.check.branch_disclosure import run_branch_checks

    root = _samples_root()
    a_pdf = root / "sample" / "光大银行2025年A股年报_含错误_测试版.pdf"
    h_pdf = root / "sample" / "光大银行_2025年H股年报.pdf"
    if not a_pdf.is_file() or not h_pdf.is_file():
        pytest.skip("CEB sample pair missing")

    diffs, diagnostics = run_branch_checks(str(a_pdf), str(h_pdf))

    assert len(diffs) == 40, f"expected 40 branch diffs, got {len(diffs)}"
    assert diagnostics["branch_diff_count"] == 40
